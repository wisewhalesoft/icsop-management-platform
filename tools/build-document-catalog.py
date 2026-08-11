# -*- coding: utf-8 -*-
"""
由 reference/程序書目錄清單(1150805).xlsx 產生
backend/src/database/seeds/document-catalog.json（匯入用中繼資料）。

執行：
    python tools/build-document-catalog.py

來源工作表為單一「整理後」表；第 1 列為小計、第 2 列為表頭、第 3 列起為資料。
本腳本只做「機械化清洗」，不做任何組織代碼推斷——組織/室長之解析留給
seed-document-catalog.ts（組織對應表為 document-catalog-org-map.json，人工可覆寫）。
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "reference" / "程序書目錄清單(1150805).xlsx"
OUT = ROOT / "backend" / "src" / "database" / "seeds" / "document-catalog.json"

# 表頭（第 2 列）→ 0-based 欄索引
COL_COMPANY, COL_DEPT, COL_SECTION, COL_CHIEF = 0, 1, 2, 3
COL_NUMBER, COL_NAME, COL_SUMMARY = 7, 8, 9
COL_CYCLE = 13

# 循環別格式：「銷售及收款循環(消金)(SRC)」／「薪工循環(LWC)」
CYCLE_RE = re.compile(r"^(?P<name>[^(（]+?)(?:[(（](?P<sub>[^)）]+)[)）])?[(（](?P<code>[A-Z]+)[)）]$")


def norm(v):
    """儲存格 → 去頭尾空白之字串；空值/空白 → None。內含定位字元一併清除（實測第 21 列編號夾 \\t）。"""
    if v is None:
        return None
    s = re.sub(r"\s+", "", str(v)) if isinstance(v, str) else str(v)
    return s or None


def norm_text(v):
    """文字欄位：僅去頭尾空白，保留內部字元。"""
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def parse_cycle(raw):
    """『循環別』→ (name, subcategory)。與 LIFECYCLE 之 (name, subcategory) 對應；尾端英文代碼捨棄。"""
    m = CYCLE_RE.match(raw.strip())
    if not m:
        raise ValueError(f"無法解析循環別：{raw!r}")
    return m.group("name").strip(), (m.group("sub").strip() if m.group("sub") else None)


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb.worksheets[0]

    records, skipped, dup = [], [], []
    seen = {}
    for r in range(3, ws.max_row + 1):
        cells = [ws.cell(r, c).value for c in range(1, 24)]
        number = norm(cells[COL_NUMBER])
        name = norm_text(cells[COL_NAME])
        if not number or not name:
            if any(x is not None for x in cells[:20]):
                skipped.append({"row": r, "reason": "缺編號或書名"})
            continue

        cycle_raw = norm_text(cells[COL_CYCLE])
        if not cycle_raw:
            skipped.append({"row": r, "reason": "缺循環別", "documentNumber": number})
            continue
        cycle_name, cycle_sub = parse_cycle(cycle_raw)

        rec = {
            "sourceRow": r,
            "documentNumber": number,
            "documentName": name,
            "contentSummary": norm_text(cells[COL_SUMMARY]),
            "lifecycleName": cycle_name,
            "lifecycleSubcategory": cycle_sub,
            "companyLabel": norm_text(cells[COL_COMPANY]),
            "deptLabel": norm_text(cells[COL_DEPT]),
            "sectionLabel": norm_text(cells[COL_SECTION]),
            "chiefName": norm_text(cells[COL_CHIEF]),
        }
        if number in seen:
            dup.append({"row": r, "documentNumber": number, "firstRow": seen[number]})
            continue
        seen[number] = r
        records.append(rec)

    payload = {
        "source": SRC.name,
        "generatedBy": "tools/build-document-catalog.py",
        "count": len(records),
        "duplicatesDropped": dup,
        "skipped": skipped,
        "records": records,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[build-document-catalog] {len(records)} 筆 → {OUT.relative_to(ROOT)}")
    print(f"  重複編號捨棄：{len(dup)}；略過列：{len(skipped)}")
    for d in dup:
        print(f"    dup  列{d['row']} {d['documentNumber']}（首見於列 {d['firstRow']}）")
    for s in skipped:
        print(f"    skip 列{s['row']} {s['reason']}")


if __name__ == "__main__":
    sys.exit(main())
